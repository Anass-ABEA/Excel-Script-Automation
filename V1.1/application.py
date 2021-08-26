import glob
import xlsxwriter
import openpyxl
from xlsxwriter.exceptions import FileCreateError

def cell(sheet, e):
    b = ord(e[0]) - ord("A") + 1
    a = int(e[1:])
    return sheet.cell(a, b).value


def getDonneesFichier(filePath):
    workbook = openpyxl.load_workbook(filePath)
    sheet = workbook.active
    result = {}
    result["entreprise"] = cell(sheet, "B3")
    result["Ville"] = cell(sheet, "B4")
    result["Tel"] = cell(sheet, "B5")
    result["Site"] = cell(sheet, "B6")
    result["RH"] = cell(sheet, "B7")
    result["RH email"] = cell(sheet, "B8")
    result["Encadrant"] = cell(sheet, "B9")
    result["Encadrant email"] = cell(sheet, "B10")
    result["Etudiant"] = cell(sheet, "B12")
    result["Niveau"] = cell(sheet, "B13")
    result["sujet"] = cell(sheet, "B15")
    result["nature Sujet"] = cell(sheet, "B16")
    result["Secteur Sujet"] = cell(sheet, "B17")
    result["Durée Sujet"] = cell(sheet, "B19")
    result["Annee Stage"] = cell(sheet, "B20")
    result["Note Stage"] = cell(sheet, "B27")
    result["Observation Stage"] = cell(sheet, "B28")
    workbook.close()
    return result



def initaliserPageEntreprise(page):
    cell_format = workbook.add_format()
    cell_format.set_bold()
    cell_format.set_align("center")
    cell_format.set_align("vcenter")

    page.set_column(0, 8, 30)

    ligne = 0
    page.write(ligne, 0, "entreprise", cell_format)
    page.write(ligne, 1, "Ville", cell_format)
    page.write(ligne, 2, "Tel", cell_format)
    page.write(ligne, 3, "Site", cell_format)
    page.write(ligne, 4, "RH", cell_format)
    page.write(ligne, 5, "RH email", cell_format)
    page.write(ligne, 6, "Encadrant", cell_format)
    page.write(ligne, 7, "Encadrant email", cell_format)
    page.write(ligne, 8, "nature Sujet", cell_format)

def initaliserPagesEtudiants(page, val):
    cell_format = workbook.add_format()
    cell_format.set_bold()
    cell_format.set_align("center")
    cell_format.set_align("vcenter")


    page.set_column(0, 8, 30)

    page.merge_range("A1:J2", "merged")
    page.write(0, 0, "Etudiants {} Année".format(val), cell_format)
    ligne = 2
    page.write(ligne, 0, "Etudiant", cell_format)
    page.write(ligne, 1, "Niveau", cell_format)
    page.write(ligne, 2, "sujet", cell_format)
    page.write(ligne, 3, "nature Sujet", cell_format)
    page.write(ligne, 4, "Secteur Sujet", cell_format)
    page.write(ligne, 5, "Durée Sujet", cell_format)
    page.write(ligne, 6, "Annee Stage", cell_format)
    page.write(ligne, 7, "Note Stage", cell_format)
    page.write(ligne, 8, "Observation Stage", cell_format)


def addToyear(page_etudiants, data, ligne):
    page_etudiants.write(ligne, 0, data["Etudiant"])
    page_etudiants.write(ligne, 1, data["Niveau"])
    page_etudiants.write(ligne, 2, data["sujet"])
    page_etudiants.write(ligne, 3, data["nature Sujet"])
    page_etudiants.write(ligne, 4, data["Secteur Sujet"])
    page_etudiants.write(ligne, 5, data["Durée Sujet"])
    page_etudiants.write(ligne, 6, data["Annee Stage"])
    page_etudiants.write(ligne, 7, data["Note Stage"])
    page_etudiants.write(ligne, 8, data["Observation Stage"])

def addToCompaniesSheet(page_entreprises, data, ligne):
    page_entreprises.write(ligne, 0, data["entreprise"])
    page_entreprises.write(ligne, 1, data["Ville"])
    page_entreprises.write(ligne, 2, data["Tel"])
    page_entreprises.write(ligne, 3, data["Site"])
    page_entreprises.write(ligne, 4, data["RH"])
    page_entreprises.write(ligne, 5, data["RH email"])
    page_entreprises.write(ligne, 6, data["Encadrant"])
    page_entreprises.write(ligne, 7, data["Encadrant email"])
    page_entreprises.write(ligne, 8, data["nature Sujet"])


# lire la liste des fichiers Excel
excel_list = glob.glob("XLSX/*.xlsx")

# creation d'un nouveau fichier excel pour y mettre le résultat
workbook = xlsxwriter.Workbook("./ContactEntreprises.xlsx")

# page entreprises et étudiants
page_entreprises = workbook.add_worksheet("entreprises")
page_etudiants1A = workbook.add_worksheet("1A")
page_etudiants2A = workbook.add_worksheet("2A")
page_etudiants3A = workbook.add_worksheet("3A")

# initialisation des pages ( ajout du header du tableau)
initaliserPagesEtudiants(page_etudiants1A, "1ère")
initaliserPagesEtudiants(page_etudiants2A, "2ème")
initaliserPagesEtudiants(page_etudiants3A, "3ème")
initaliserPageEntreprise(page_entreprises)

# Emails RH visités ( éviter la duplication )
rh_encadrantList = []
listeEntreprises = []
ligne_1A = 3
ligne_2A = 3
ligne_3A = 3
ligne_S = 2

try:
    for fichier in excel_list:
        data = getDonneesFichier(fichier)
        if (data["RH email"],data["Encadrant email"]) not in rh_encadrantList:
            addToCompaniesSheet(page_entreprises, data, len(rh_encadrantList) + 1)
            rh_encadrantList.append((data["RH email"], data["Encadrant email"]))
        if data["Niveau"] == "1ère année":
            addToyear(page_etudiants1A, data, ligne_1A)
            ligne_1A += 1
        elif data["Niveau"] == "2ème Année":
            addToyear(page_etudiants2A, data, ligne_2A)
            ligne_2A += 1
        else:
            addToyear(page_etudiants3A, data, ligne_3A)
            ligne_3A += 1

    workbook.close()

except FileCreateError as err:
    print("Veuillez fermer les fichiers Excels:\n\t• dans le dossier XLSX (si ouvert)\n\t• 'ContactEntreprises.xlsx' (si ouvert)\nPuis relancez l'application")